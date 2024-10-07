from django.shortcuts import render, redirect, get_object_or_404
from .forms import CourseForm, ClassForm, AddShiftForm, ScheduleForm, CourseOfferingForm, TeacherForm, addday, addroom, addslot, CourseAssignmentForm, DepartmentForm, TeacherAssignmentForm
from .models import Room, Timeslot,Shift, Day, Schedule, Class, Teacher, Course, CourseAssignment, Department,  CourseOffering, Semester
from django.http import JsonResponse, HttpResponse
import openpyxl


def timetable_view(request):
    days = Day.objects.all()
    rooms = Room.objects.all()
    timeslots = Timeslot.objects.all()

    schedules = Schedule.objects.all()

    context = {
        'days': days,
        'rooms': rooms,
        'timeslots': timeslots,
        'schedules': schedules,
    }
    return render(request, 'timetable/timetable.html', context)


def add_course(request):
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save(commit=False)
            
            # Check if a course with the same full name already exists
            if Course.objects.filter(full_name=course.full_name).exists():
                # Add non-field error to form
                form.add_error(None, "Course with this full name already exists.")
            elif Course.objects.filter(course_code=course.course_code).exists():
                # Add non-field error if the course code already exists
                form.add_error(None, "Course with this course code already exists.")
            else:
                # Save the new course
                course.save()
                return redirect('dashboard')  # Or redirect to the appropriate page
    else:
        form = CourseForm()
    
    return render(request, 'timetable/add_course.html', {'form': form})

def add_class(request):
    if request.method == 'POST':
        form = ClassForm(request.POST)
        if form.is_valid():
            class_instance = form.save(commit=False)
            
            # **New: Get or create the Shift instance**
            shift_name = request.POST.get('shift')
            try:
                shift_instance = Shift.objects.get(name=shift_name)
            except Shift.DoesNotExist:
                return HttpResponse(f"Shift {shift_name} does not exist.", status=400)

            class_instance.shift = shift_instance  # **Assign the shift instance**

            # Check if the class already exists in the department and semester
            if Class.objects.filter(name=class_instance.name, section=class_instance.section, department=class_instance.department, semester=class_instance.semester, shift=class_instance.shift).exists():
                return HttpResponse("Class already exists in this department, semester, and shift.", status=400)
            
            class_instance.save()
            return redirect('dashboard')
    else:
        form = ClassForm()
    
    return render(request, 'timetable/add_class.html', {'form': form})



def assign_course_to_class(request):
    form = CourseAssignmentForm()

    if request.method == 'POST':
        department_id = request.POST.get('department')
        semester_id = request.POST.get('semester')
        
        form = CourseAssignmentForm(request.POST, department_id=department_id, semester_id=semester_id)
        
        if form.is_valid():
            selected_class = form.cleaned_data['class_assigned']
            selected_courses = form.cleaned_data['courses']

            for course in selected_courses:
                CourseAssignment.objects.get_or_create(
                    course=course,
                    class_assigned=selected_class
                )

            return redirect('assign_teacher_to_course')

    return render(request, 'timetable/assign_course_to_class.html', {'form': form})

def load_classes_and_courses(request):
    department_id = request.GET.get('department')
    semester_id = request.GET.get('semester')
    
    if department_id and semester_id:
        # Get classes with combined information
        classes = list(
            Class.objects.filter(department_id=department_id, semester_id=semester_id)
            .values('id', 'name', 'semester__name', 'section')
        )
        
        # Format class info
        formatted_classes = [
            {
                'id': cls['id'],
                'class_info': f"{cls['name']} {cls['semester__name']}{cls['section']}"
            }
            for cls in classes
        ]

        # Get courses with distinct offerings
        courses = list(
            Course.objects.filter(courseoffering__department_id=department_id, courseoffering__semester_id=semester_id)
            .distinct()
            .values('id', 'full_name')
        )
    else:
        formatted_classes = []
        courses = []
    
    return JsonResponse({'classes': formatted_classes, 'courses': courses})


def get_courses(request, department_id):
    semester_id = request.GET.get('semester_id')
    course_offerings = CourseOffering.objects.filter(department_id=department_id, semester_id=semester_id)
    courses = [offering.course for offering in course_offerings]
    return render(request, 'course_list.html', {'courses': courses})

def get_classes(request, department_id):
    semester_id = request.GET.get('semester_id')
    classes = Class.objects.filter(department_id=department_id, semester_id=semester_id)
    return render(request, 'class_list.html', {'classes': classes})

def assign_teacher_to_course(request):
    if request.method == 'POST':
        form = TeacherAssignmentForm(request.POST)
        if form.is_valid():
            course_assignment = form.cleaned_data['course_assignment']
            teacher = form.cleaned_data['teacher']

            course_assignment.teacher = teacher
            course_assignment.save()

            # Success message after creation
            return render(request, 'assign_teacher_to_course.html', {
            'success_message': f"Successfully Teacher Assigned." })
    else:
        form = TeacherAssignmentForm()

    return render(request, 'timetable/assign_teacher_to_course.html', {'form': form})

def load_course_class_pairs(request):
    department_id = request.GET.get('department')
    semester_id = request.GET.get('semester')

    if department_id and semester_id:
        # Fetch course assignments based on department and semester
        course_assignments = CourseAssignment.objects.filter(
            class_assigned__department_id=department_id,
            class_assigned__semester_id=semester_id
        ).select_related('class_assigned', 'course')

        # Prepare the response data
        response_data = []
        for assignment in course_assignments:
            response_data.append({
                'id': assignment.id,
                'course__short_name': assignment.course.short_name,
                'class_info': f"{assignment.class_assigned.name} {assignment.class_assigned.semester.name}{assignment.class_assigned.section}"
            })
    else:
        response_data = []

    # Return the JSON response
    return JsonResponse({'course_assignments': response_data})

def get_course_assignments(request, department_id):
    course_assignments = CourseAssignment.objects.filter(class_assigned__department_id=department_id)
    data = [{'id': ca.id, 'course_name': ca.course.name, 'class_name': ca.class_assigned.name} for ca in course_assignments]
    return JsonResponse({'course_assignments': data})

def assign_schedule(request):
    # Get departments and semesters for the form
    departments = Department.objects.all()
    semesters = Semester.objects.all()

    if request.method == 'POST':
        department_id = request.POST.get('department')
        semester_id = request.POST.get('semester')
        class_id = request.POST.get('class_select')
        course_assignment_id = request.POST.get('course_assignment')

        # Initialize the form with POST data and extra arguments
        form = ScheduleForm(request.POST, department_id=department_id, semester_id=semester_id, class_id=class_id)
        
        if form.is_valid():
            day = form.cleaned_data['day']
            room = form.cleaned_data['room']
            timeslot = form.cleaned_data['timeslot']
            department = form.cleaned_data['department']
            semester = form.cleaned_data['semester']
            class_assigned = form.cleaned_data['class_select']
            
            try:
                # Retrieve CourseAssignment object using the provided ID
                course_assign = CourseAssignment.objects.get(id=course_assignment_id)
            except CourseAssignment.DoesNotExist:
                form.add_error(None, "Selected course assignment does not exist.")
                return render(request, 'timetable/assign_schedule.html', {
                    'form': form,
                    'departments': departments,
                    'semesters': semesters
                })

            # Check room, teacher, and class availability
            if Schedule.objects.filter(day=day, room=room, timeslot=timeslot).exists():
                form.add_error(None, "This room is already booked for the selected timeslot.")
            elif Schedule.objects.filter(day=day, timeslot=timeslot, course_assignment__teacher=course_assign.teacher).exclude(course_assignment__course=course_assign.course).exists():
                form.add_error(None, "This teacher is already assigned to another course at this time.")
            elif Schedule.objects.filter(day=day, timeslot=timeslot, course_assignment__class_assigned=course_assign.class_assigned).exists():
                form.add_error(None, "This class is already assigned to a different room at this time.")
            else:
                # All checks passed, save the schedule
                Schedule.objects.create(day=day, room=room, timeslot=timeslot, course_assignment=course_assign)
                return redirect('timetable')  # Redirect to the timetable page
        else:
            # Add form errors if validation fails
            form.add_error(None, "Form submission error. Please check your inputs.")
    else:
        form = ScheduleForm()

    return render(request, 'timetable/assign_schedule.html', {
        'form': form,
        'departments': departments,
        'semesters': semesters
    })


def load_classes(request):
    department_id = request.GET.get('department_id')
    semester_id = request.GET.get('semester_id')

    if department_id and semester_id:
        classes = Class.objects.filter(department_id=department_id, semester_id=semester_id)
        class_list = list(classes.values('id', 'name', 'section'))
    else:
        class_list = []

    return JsonResponse(class_list, safe=False)

def load_course_assignments(request):
    class_id = request.GET.get('class_id')

    if class_id:
        assignments = CourseAssignment.objects.filter(class_assigned_id=class_id)
        assignment_list = list(assignments.values('id', 'course__full_name'))
    else:
        assignment_list = []

    return JsonResponse(assignment_list, safe=False)

def load_classes_by_department_and_semester(request):
    department_id = request.GET.get('department_id')
    semester_id = request.GET.get('semester_id')
    classes = Class.objects.filter(department_id=department_id, semester_id=semester_id)
    class_list = list(classes.values('id', 'name', 'section'))
    return JsonResponse(class_list, safe=False)


def load_course_assignments_by_class(request):
    class_id = request.GET.get('class_id')
    assignments = CourseAssignment.objects.filter(class_assigned_id=class_id)
    assignment_list = list(assignments.values('id', 'course__full_name'))
    return JsonResponse(assignment_list, safe=False)


def get_class_list(request):
    department_id = request.GET.get('department')
    semester_id = request.GET.get('semester')

    classes = Class.objects.filter(department_id=department_id, semester_id=semester_id).values('id', 'name')
    return JsonResponse(list(classes), safe=False)
def get_course_assignments(request):
    class_id = request.GET.get('class_id')

    course_assignments = CourseAssignment.objects.filter(class_assigned_id=class_id).values('id', 'course__full_name')
    data = [{'id': ca['id'], 'course_name': ca['course__full_name']} for ca in course_assignments]
    return JsonResponse(data, safe=False)

def filter_classes(request):
    department_id = request.GET.get('department')
    semester_id = request.GET.get('semester')

    if department_id and semester_id:
        classes = Class.objects.filter(department_id=department_id, semester_id=semester_id).values('id', 'name')
        return JsonResponse(list(classes), safe=False)
    else:
        return JsonResponse([], safe=False)

def filter_course_assignments(request):
    class_id = request.GET.get('class_id')

    if class_id:
        course_assignments = CourseAssignment.objects.filter(class_assigned_id=class_id).values('id', 'course__full_name')
        return JsonResponse(list(course_assignments), safe=False)
    return JsonResponse([], safe=False)
def add_teacher(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST)
        if form.is_valid():
            form.save()
            # Success message after creation
        return render(request, 'timetable/add_teacher.html', {
            'success_message': f"Successfully Add Teacher." })
    else:
        form = TeacherForm()

    return render(request, 'timetable/add_teacher.html', {'form': form})


def class_timetable_selection_view(request):
    departments = Department.objects.all()  # Get all departments
    selected_department_id = request.GET.get('department_id')
    selected_class_id = request.GET.get('class_id')

    classes = Class.objects.none()  # Default to empty queryset

    # Convert selected_department_id to an integer if it's provided
    if selected_department_id:
        selected_department_id = int(selected_department_id)

    if selected_department_id:
        classes = Class.objects.filter(department_id=selected_department_id)  # Filter classes by department

    if selected_class_id:
        return redirect('class_timetable', class_id=selected_class_id)

    context = {
        'departments': departments,
        'classes': classes,
        'selected_department_id': selected_department_id,  # Ensure this is passed to the template
    }

    return render(request, 'timetable/class_timetable_selection.html', context)


def class_timetable_view(request, class_id):
    selected_class = Class.objects.get(id=class_id)
    selected_department_id = selected_class.department.id  # Get the department of the selected class
    days = Day.objects.all()
    timeslots = Timeslot.objects.all()

    # Prepare timetable data in a more template-friendly format
    timetable_data = []
    for day in days:
        day_row = {'day': day.name, 'slots': []}
        for timeslot in timeslots:
            schedules = Schedule.objects.filter(
                course_assignment__class_assigned=selected_class,
                day=day,
                timeslot=timeslot
            )
            if schedules.exists():
                slot_data = []
                for schedule in schedules:
                    slot_data.append({
                        'room_name': schedule.room.name,
                        'course_name': schedule.course_assignment.course.short_name,  # Use short name
                        'teacher_name': schedule.course_assignment.teacher.name if schedule.course_assignment.teacher else 'TBA'
                    })
                day_row['slots'].append(slot_data)
            else:
                day_row['slots'].append([])  # Empty slot
        timetable_data.append(day_row)

    context = {
        'selected_class': selected_class,
        'days': days,
        'timeslots': timeslots,
        'timetable_data': timetable_data,
        'selected_department_id': selected_department_id,  # Add the selected department
    }
    return render(request, 'timetable/class_timetable.html', context)


def teacher_timetable_selection_view(request):
    teachers = Teacher.objects.all()
    selected_teacher_id = request.GET.get('teacher_id')
    if selected_teacher_id:
        return redirect('teacher_timetable', teacher_id=selected_teacher_id)
    return render(request, 'timetable/teacher_timetable_selection.html', {'teachers': teachers})

def teacher_timetable_view(request, teacher_id):
    selected_teacher = Teacher.objects.get(id=teacher_id)
    days = Day.objects.all()
    timeslots = Timeslot.objects.all()

    # Prepare timetable data in a more template-friendly format
    timetable_data = []
    for day in days:
        day_row = {'day': day.name, 'slots': []}
        for timeslot in timeslots:
            # Filter schedules by teacher, day, and timeslot
            schedules = Schedule.objects.filter(course_assignment__teacher_id=teacher_id, day=day, timeslot=timeslot)
            if schedules.exists():
                schedule_info = []
                for schedule in schedules:
                    schedule_info.append({
                        'room_name': schedule.room.name,
                        'course_name': schedule.course_assignment.course.short_name if schedule.course_assignment.course else 'N/A',
                        'class_name': f"{schedule.course_assignment.class_assigned.name} {schedule.course_assignment.class_assigned.semester.name}{schedule.course_assignment.class_assigned.section}"
                    })
                day_row['slots'].append(schedule_info)
            else:
                day_row['slots'].append([])  # Empty slot
        timetable_data.append(day_row)

    context = {
        'selected_teacher': selected_teacher,
        'days': days,
        'timeslots': timeslots,
        'timetable_data': timetable_data
    }
    return render(request, 'timetable/teacher_timetable.html', context)

def dashboard_view(request):
    return render(request, 'timetable/dashboard.html')

def add_data(request):
    return render(request, 'timetable/add_data.html')

def full_timetable(request):
    days = Day.objects.all()
    timeslots = Timeslot.objects.all()
    rooms = Room.objects.all()
    schedules = Schedule.objects.select_related('assigned_class', 'course', 'teacher', 'room', 'timeslot', 'day').all()

    return render(request, 'timetable/full_timetable.html', {
        'days': days,
        'timeslots': timeslots,
        'rooms': rooms,
        'schedules': schedules
    })

def teacher_contacts(request):
    teachers = Teacher.objects.all()

    context = {
        'teachers': teachers,
    }
    return render(request, 'timetable/teacher_contacts.html', context)

def Courses_details(request):
    Courses = Course.objects.all()

    context = {
        'Courses': Courses,
    }
    return render(request, 'timetable/Courses_details.html', context)


def delete_schedule(request):
    error_message = None
    success_message = None
    schedule_info = None  # Variable to hold the schedule details

    if request.method == 'POST':
        if 'check_schedule' in request.POST:
            day_id = request.POST.get('day')
            room_id = request.POST.get('room')
            timeslot_id = request.POST.get('timeslot')

            if not day_id or not room_id or not timeslot_id:
                error_message = "Please select day, room, and timeslot."
            else:
                try:
                    # Fetch the schedule for the selected day, room, and timeslot
                    schedule_info = Schedule.objects.get(day_id=day_id, room_id=room_id, timeslot_id=timeslot_id)
                except Schedule.DoesNotExist:
                    error_message = "No schedule exists for the selected day, room, and timeslot."
        
        elif 'confirm_delete' in request.POST:
            day_id = request.POST.get('day')
            room_id = request.POST.get('room')
            timeslot_id = request.POST.get('timeslot')

            if not day_id or not room_id or not timeslot_id:
                error_message = "Please select day, room, and timeslot."
            else:
                try:
                    schedule = Schedule.objects.get(day_id=day_id, room_id=room_id, timeslot_id=timeslot_id)
                    schedule.delete()  # Delete the schedule
                    success_message = "Schedule deleted successfully."
                except Schedule.DoesNotExist:
                    error_message = "No schedule exists for the selected day, room, and timeslot."

    days = Day.objects.all()
    rooms = Room.objects.all()
    timeslots = Timeslot.objects.all()

    context = {
        'days': days,
        'rooms': rooms,
        'timeslots': timeslots,
        'schedule_info': schedule_info,
        'error_message': error_message,
        'success_message': success_message,
    }
    return render(request, 'timetable/delete_schedule.html', context)



from django.db import transaction
from django.shortcuts import render, redirect
from .forms import ScheduleSwapForm
from .models import Schedule, Room, Timeslot, Day

from django import forms  # Import this if you're using forms for validation

def swap_schedules(request):
    schedule_info_a = None
    schedule_info_b = None
    error_message = None
    success_message = None

    if request.method == 'POST':
        if 'check_schedules' in request.POST:
            day_a = request.POST.get('day_a')
            room_a = request.POST.get('room_a')
            timeslot_a = request.POST.get('timeslot_a')

            day_b = request.POST.get('day_b')
            room_b = request.POST.get('room_b')
            timeslot_b = request.POST.get('timeslot_b')

            if not day_a or not room_a or not timeslot_a or not day_b or not room_b or not timeslot_b:
                error_message = "Please select both sets of day, room, and timeslot."
            else:
                try:
                    schedule_info_a = Schedule.objects.get(day_id=day_a, room_id=room_a, timeslot_id=timeslot_a)
                    schedule_info_b = Schedule.objects.get(day_id=day_b, room_id=room_b, timeslot_id=timeslot_b)
                except Schedule.DoesNotExist:
                    error_message = "One or both of the schedules do not exist."

        elif 'confirm_swap' in request.POST:
            day_a = request.POST.get('day_a')
            room_a = request.POST.get('room_a')
            timeslot_a = request.POST.get('timeslot_a')

            day_b = request.POST.get('day_b')
            room_b = request.POST.get('room_b')
            timeslot_b = request.POST.get('timeslot_b')

            try:
                # Fetch the actual instances of Day, Room, and Timeslot
                day_a_instance = Day.objects.get(id=day_a)
                room_a_instance = Room.objects.get(id=room_a)
                timeslot_a_instance = Timeslot.objects.get(id=timeslot_a)

                day_b_instance = Day.objects.get(id=day_b)
                room_b_instance = Room.objects.get(id=room_b)
                timeslot_b_instance = Timeslot.objects.get(id=timeslot_b)

                schedule_a = Schedule.objects.get(day=day_a_instance, room=room_a_instance, timeslot=timeslot_a_instance)
                schedule_b = Schedule.objects.get(day=day_b_instance, room=room_b_instance, timeslot=timeslot_b_instance)

                course_assign_a = schedule_a.course_assignment
                course_assign_b = schedule_b.course_assignment

                # Validation checks
                # Check teacher availability for Schedule A
                if Schedule.objects.filter(
                        day=day_b_instance, 
                        timeslot=timeslot_b_instance, 
                        course_assignment__teacher=course_assign_a.teacher
                    ).exclude(course_assignment__course=course_assign_a.course).exists():
                    error_message = "Teacher for Schedule A is already assigned to another course at this time."

                # Check teacher availability for Schedule B
                elif Schedule.objects.filter(
                        day=day_a_instance, 
                        timeslot=timeslot_a_instance, 
                        course_assignment__teacher=course_assign_b.teacher
                    ).exclude(course_assignment__course=course_assign_b.course).exists():
                    error_message = "Teacher for Schedule B is already assigned to another course at this time."

                # Check class availability for Schedule A
                elif Schedule.objects.filter(
                        day=day_b_instance, 
                        timeslot=timeslot_b_instance, 
                        course_assignment__class_assigned=course_assign_a.class_assigned
                    ).exists():
                    error_message = "Class for Schedule A is already assigned to a different room at this time."

                # Check class availability for Schedule B
                elif Schedule.objects.filter(
                        day=day_a_instance, 
                        timeslot=timeslot_a_instance, 
                        course_assignment__class_assigned=course_assign_b.class_assigned
                    ).exists():
                    error_message = "Class for Schedule B is already assigned to a different room at this time."

                # If no errors, proceed with the swap
                else:
                    with transaction.atomic():
                        # Temporary room and timeslot for swapping
                        temp_room = Room.objects.create(name="TEMP_ROOM")
                        temp_timeslot = Timeslot.objects.create(slot="TEMP_SLOT")

                        # Move schedule A to temporary slot
                        schedule_a.room = temp_room
                        schedule_a.timeslot = temp_timeslot
                        schedule_a.save()

                        # Move schedule B to A's original slot
                        schedule_b.day = day_a_instance
                        schedule_b.room = room_a_instance
                        schedule_b.timeslot = timeslot_a_instance
                        schedule_b.save()

                        # Move schedule A to B's original slot
                        schedule_a.day = day_b_instance
                        schedule_a.room = room_b_instance
                        schedule_a.timeslot = timeslot_b_instance
                        schedule_a.save()

                        # Delete temp room and timeslot
                        temp_room.delete()
                        temp_timeslot.delete()

                    success_message = "Schedules swapped successfully."

            except Schedule.DoesNotExist:
                error_message = "One or both of the schedules do not exist."

    days = Day.objects.all()
    rooms = Room.objects.all()
    timeslots = Timeslot.objects.all()

    context = {
        'days': days,
        'rooms': rooms,
        'timeslots': timeslots,
        'schedule_info_a': schedule_info_a,
        'schedule_info_b': schedule_info_b,
        'error_message': error_message,
        'success_message': success_message,
    }

    return render(request, 'timetable/swap_schedules.html', context)



from django.http import HttpResponseBadRequest
def add_course_offering(request):
    if request.method == 'POST':
        course_id = request.POST.get('course_id')
        if not course_id:
            return HttpResponseBadRequest("Course ID is missing")
        
        course = get_object_or_404(Course, id=course_id)
        
        # Retrieve all department-semester-year combinations
        department_ids = request.POST.getlist('department')
        semester_ids = request.POST.getlist('semester')
        year = request.POST.get('year')  # Get the year from the form

        if not department_ids or not semester_ids or not year:
            return HttpResponseBadRequest("Please select at least one department, semester, and provide the year.")

        # Loop through each department-semester pair
        for department_id, semester_id in zip(department_ids, semester_ids):
            try:
                department = get_object_or_404(Department, id=department_id)
                semester = get_object_or_404(Semester, id=semester_id)

                # Check for duplicates
                if not CourseOffering.objects.filter(course=course, department=department, semester=semester, year=year).exists():
                    CourseOffering.objects.create(course=course, department=department, semester=semester, year=year)
            except ValueError:
                return HttpResponseBadRequest("Invalid department or semester ID")

        # Success message after creation
        return render(request, 'timetable/add_course_offering.html', {
            'success_message': f"Successfully Course Offer." })

    departments = Department.objects.all()
    semesters = Semester.objects.all()

    return render(request, 'timetable/add_course_offering.html', {
        'departments': departments,
        'semesters': semesters
    })




def course_search(request):
    term = request.GET.get('term', '')
    # Query the courses based on the search term
    courses = Course.objects.filter(full_name__icontains=term).values('id', 'full_name')
    # Return the courses as JSON
    return JsonResponse(list(courses), safe=False)

def course_details(request):
    course_id = request.GET.get('course_id', '')
    course_offerings = CourseOffering.objects.filter(course_id=course_id)

    departments = Department.objects.all().values('id', 'name')
    semesters = Semester.objects.all().values('id', 'name')

    data = {
        "departments": list(departments),
        "semesters": list(semesters)
    }
    return JsonResponse(data)

def room_timetable_selection_view(request):
    rooms = Room.objects.all()
    selected_room_id = request.GET.get('room_id')
    if selected_room_id:
        return redirect('room_timetable_view', room_id=selected_room_id)
    return render(request, 'timetable/room_timetable_selection.html', {'rooms': rooms})

def room_timetable_view(request, room_id):
    selected_room = Room.objects.get(id=room_id)
    days = Day.objects.all()
    timeslots = Timeslot.objects.all()

    # Prepare timetable data in a more template-friendly format
    timetable_data = []
    for day in days:
        day_row = {'day': day.name, 'slots': []}
        for timeslot in timeslots:
            # Filter schedules by room, day, and timeslot
            schedules = Schedule.objects.filter(room_id=room_id, day=day, timeslot=timeslot)
            if schedules.exists():
                schedule_info = []
                for schedule in schedules:
                    schedule_info.append({
                        'course_name': schedule.course_assignment.course.short_name if schedule.course_assignment.course else 'N/A',
                        'class_name': f"{schedule.course_assignment.class_assigned.name} {schedule.course_assignment.class_assigned.semester.name} {schedule.course_assignment.class_assigned.section}",
                        'teacher_name': schedule.course_assignment.teacher.name if schedule.course_assignment.teacher else 'N/A'
                    })
                day_row['slots'].append(schedule_info)
            else:
                day_row['slots'].append([])  # Empty slot
        timetable_data.append(day_row)

    context = {
        'selected_room': selected_room,
        'days': days,
        'timeslots': timeslots,
        'timetable_data': timetable_data
    }
    return render(request, 'timetable/room_timetable.html', context)

from django.core.files.storage import FileSystemStorage
import pandas as pd
from openpyxl import load_workbook

def upload_courses_excel(request):
    if request.method == 'POST':
        # Check if the file is uploaded
        if not request.FILES.get('file'):
            return render(request, 'timetable/upload_courses_excel.html', {'error': 'Please select a file to upload.'})
        
        excel_file = request.FILES['file']
        
        try:
            # Load the workbook and get the active sheet
            wb = load_workbook(excel_file, read_only=True)
            sheet = wb.active  # Get the active sheet
            
            # Validate the required columns (header row)
            header = [cell.value for cell in sheet[1]]  # First row is the header
            required_columns = ['Course_code', 'Short_Name', 'Full_Name', 'credit_hours', 'lab_crh', 'pre_req', 'Category']
            
            if not all(col in header for col in required_columns):
                return render(request, 'timetable/upload_courses_excel.html', {'error': 'Missing required columns in Excel file.'})
            
            # Find indices of the required columns
            col_indices = {col: header.index(col) for col in required_columns}
            
            # List to collect courses to be bulk created
            courses_to_create = []
            row_count = 0
            
            # Iterate through the rows, starting from row 2 (skip header)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_count += 1
                
                # Extract course data from the current row
                course_data = {
                    'course_code': row[col_indices['Course_code']],
                    'short_name': row[col_indices['Short_Name']],
                    'full_name': row[col_indices['Full_Name']],
                    'credit_hours': row[col_indices['credit_hours']],
                    'lab_crh': row[col_indices['lab_crh']],
                    'pre_req': row[col_indices['pre_req']],
                    'category': row[col_indices['Category']]
                }

                # Skip rows with missing course code or full name
                if not course_data['course_code'] or not course_data['full_name']:
                    continue

                # Create a Course object
                course = Course(
                    course_code=course_data['course_code'],
                    short_name=course_data['short_name'],
                    full_name=course_data['full_name'],
                    credit_hours=course_data['credit_hours'],
                    lab_crh=course_data['lab_crh'] if course_data['lab_crh'] else 0,  # Handle empty lab_crh
                    pre_req=course_data['pre_req'],  # This can be None if the field is optional
                    category=course_data['category']
                )
                
                courses_to_create.append(course)

                # Perform bulk create in batches of 1000 rows
                if len(courses_to_create) >= 1000:
                    Course.objects.bulk_create(courses_to_create, ignore_conflicts=True)
                    courses_to_create = []  # Reset list for next batch
            
            # Insert any remaining courses
            if courses_to_create:
                Course.objects.bulk_create(courses_to_create, ignore_conflicts=True)
            
            # Success message after creation
            return render(request, 'timetable/upload_courses_excel.html', {
            'success_message': f"Successfully added {row_count} Course(s)." })

        except Exception as e:
            return render(request, 'timetable/upload_courses_excel.html', {'error': f'Error processing file: {str(e)}'})

    return render(request, 'timetable/upload_courses_excel.html')


def upload_classes_excel(request):
    if request.method == 'POST' and request.FILES['file']:
        # Handle file upload directly without saving
        excel_file = request.FILES['file']

        try:
            # Load the workbook directly from the uploaded file
            wb = load_workbook(excel_file, read_only=True)
            sheet = wb.active  # Get the active sheet
        except Exception as e:
            return render(request, 'timetable/upload_classes_excel.html', {'error': 'Invalid Excel file format.'})

        # Validate the required columns (header row)
        header = [cell.value for cell in sheet[1]]  # Assuming first row is the header
        required_columns = ['Class_Name', 'Class_semster', 'Class_Section', 'Department_Name', 'Shift']

        if not all(col in header for col in required_columns):
            return render(request, 'timetable/upload_classes_excel.html', {'error': 'Missing required columns in Excel file.'})

        # Find the indices of the required columns
        col_indices = {col: header.index(col) for col in required_columns}
        
        classes_to_create = []
        row_count = 0
        
        # Iterate through the rows, starting from row 2 to skip the header
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_count += 1

            class_data = {
                'class_name': row[col_indices['Class_Name']],
                'class_semester': row[col_indices['Class_semster']],
                'class_section': row[col_indices['Class_Section']],
                'department_name': row[col_indices['Department_Name']],
                'shift_name': row[col_indices['Shift']]
            }

            # Validate row data
            if None in class_data.values():
                continue  # Skip rows with missing data

            # Get or create the Department
            department, _ = Department.objects.get_or_create(name=class_data['department_name'])

            # Get or create the Semester based on the Class_semester value
            try:
                semester = Semester.objects.get(name=class_data['class_semester'])
            except Semester.DoesNotExist:
                return render(request, 'timetable/upload_classes_excel.html', {'error': f"Semester {class_data['class_semester']} does not exist."})

            # Get or create the Shift instance
            shift_instance, _ = Shift.objects.get_or_create(name=class_data['shift_name'])

            # Create a Class instance and append it to the list for bulk creation
            class_instance = Class(
                name=class_data['class_name'],
                section=class_data['class_section'],
                department=department,
                semester=semester,
                shift=shift_instance
            )
            classes_to_create.append(class_instance)

            # Bulk create in batches (e.g., every 1000 rows)
            if len(classes_to_create) >= 1000:
                Class.objects.bulk_create(classes_to_create, ignore_conflicts=True)
                classes_to_create = []  # Reset the list for the next batch
            
        # Insert remaining rows that didn’t form a full batch
        if classes_to_create:
            Class.objects.bulk_create(classes_to_create, ignore_conflicts=True)

        # Success message after creation
        return render(request, 'timetable/upload_classes_excel.html', {
            'success_message': f"Successfully added {row_count} Class(es)."
        })

    return render(request, 'timetable/upload_classes_excel.html')
import re
def upload_teachers_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        try:
            wb = load_workbook(excel_file, data_only=True)  # Load workbook
            sheet = wb.active
        except Exception as e:
            return render(request, 'timetable/upload_teachers_excel.html', {'errors': [f"Error reading Excel file: {str(e)}"]})

        errors = []
        created_count_teacher = 0

        # Define the expected column names (ignore case)
        expected_columns = {'name', 'phone_number', 'gmail'}
        actual_columns = {str(cell.value).strip().lower() for cell in sheet[1]}  # Convert headers to lowercase

        # Check if the expected columns are in the actual columns
        if not expected_columns.issubset(actual_columns):
            return render(request, 'timetable/upload_teachers_excel.html', {
                'errors': ["Columns must exactly match 'name', 'phone_number', and 'gmail', case-insensitively."]
            })

        # Regex to validate the phone number format after cleaning
        phone_number_pattern = re.compile(r"^0\d{3}-\d{7}$")

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Map each column to its corresponding value based on header names
            row_data = dict(zip([str(cell.value).strip().lower() for cell in sheet[1]], row))

            # Extract name, phone_number, and gmail from the row
            name = row_data.get('name')
            phone_number = row_data.get('phone_number')
            gmail = row_data.get('gmail')

            # Skip rows where name is None or empty
            if not name:
                errors.append(f"Row {idx}: Missing name.")
                continue

            if phone_number:
                # Clean phone number by removing non-digit characters
                cleaned_phone_number = re.sub(r'\D', '', str(phone_number).strip())  # Remove non-digits

                # Ensure the cleaned phone number is exactly 11 digits (for proper formatting)
                if len(cleaned_phone_number) == 11 and cleaned_phone_number.startswith('0'):
                    formatted_phone_number = f"0{cleaned_phone_number[1:4]}-{cleaned_phone_number[4:]}"
                elif len(cleaned_phone_number) == 10:
                    formatted_phone_number = f"0{cleaned_phone_number[:3]}-{cleaned_phone_number[3:]}"
                else:
                    errors.append(f"Row {idx}: Invalid phone number length or format: {phone_number}.")
                    continue

                # Validate the formatted phone number
                if not phone_number_pattern.match(formatted_phone_number):
                    errors.append(f"Row {idx}: Invalid phone number format: {formatted_phone_number}. Expected format: 0###-#######.")
                    continue
            else:
                formatted_phone_number = None  # Handle optional phone numbers (None)

            # If teacher doesn't already exist, create a new one
            if not Teacher.objects.filter(name=name, phone_number=formatted_phone_number).exists():
                Teacher.objects.create(name=name, phone_number=formatted_phone_number, gmail=gmail)
                created_count_teacher += 1
            else:
                errors.append(f"Row {idx}: Teacher with name '{name}' and phone number '{formatted_phone_number}' already exists.")

        # Display errors in the template if any
        if errors:
            return render(request, 'timetable/upload_teachers_excel.html', {'errors': errors})

        # Success message after creation
        return render(request, 'timetable/upload_teachers_excel.html', {
            'success_message': f"Successfully added {created_count_teacher} teacher(s)."
        })

    return render(request, 'timetable/upload_teachers_excel.html')

def upload_course_offerings(request):
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        
        # Attempt to read the Excel file
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
        except Exception as e:
            return render(request, 'timetable/upload_course_offerings.html', {'errors': [f"Error reading Excel file: {str(e)}"]})

        errors = []
        created_count = 0  # Count of successfully created CourseOffering instances

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Expecting 5 columns
            if len(row) != 5:
                errors.append(f"Row {idx}: Invalid number of columns.")
                continue

            course_code, course_short_name, department_name, semester_name, year = row

            # Print debug info
            print(f"Processing Row {idx}: {row}")

            # Validate year
            if not year or not isinstance(year, int):
                errors.append(f"Row {idx}: Invalid year.")
                continue

            # Validate department
            try:
                department = Department.objects.get(name=department_name)
            except Department.DoesNotExist:
                errors.append(f"Row {idx}: Department '{department_name}' does not exist.")
                continue

            # Validate semester
            try:
                semester = Semester.objects.get(name=semester_name)
            except Semester.DoesNotExist:
                errors.append(f"Row {idx}: Semester '{semester_name}' does not exist.")
                continue

            # Validate course using both course_code and course_short_name
            try:
                course = Course.objects.get(course_code=course_code, short_name=course_short_name)
            except Course.DoesNotExist:
                errors.append(f"Row {idx}: Course with code '{course_code}' and short name '{course_short_name}' does not exist.")
                continue
            
            # Create CourseOffering if it doesn't already exist
            if not CourseOffering.objects.filter(course=course, department=department, semester=semester, year=year).exists():
                CourseOffering.objects.create(course=course, department=department, semester=semester, year=year)
                created_count += 1

        # Prepare response data
        if errors:
            return render(request, 'timetable/upload_course_offerings.html', {'errors': errors})

        return render(request, 'timetable/upload_course_offerings.html', {
            'success_message': f"Successfully created {created_count} Course Offering(s)."
        })

    return render(request, 'timetable/upload_course_offerings.html')

def upload_course_assignments_excel(request):
    errors = []  # Initialize a list to store error messages

    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            errors.append('Invalid Excel file format.')
            return render(request, 'upload_course_assignments.html', {'errors': errors})

        for index, row in df.iterrows():
            class_name = row.get('Class_Name')
            semester_name = row.get('Class_semster')
            section_name = row.get('Class_Section')
            course_code = row.get('Course_code')
            teacher_name = row.get('Teacher_name')

            # Skip rows with missing required fields
            if pd.isna(class_name) or pd.isna(semester_name) or pd.isna(section_name) or pd.isna(course_code):
                continue

            semester_name = str(semester_name).strip()

            try:
                class_assigned = Class.objects.get(
                    name__iexact=str(class_name).strip(),
                    section__iexact=str(section_name).strip(),
                    semester__name__iexact=semester_name
                )
            except Class.DoesNotExist:
                errors.append(f"Row {index + 1}: Class '{class_name}', Semester '{semester_name}', Section '{section_name}' does not exist.")
                continue

            try:
                course = Course.objects.get(course_code=str(course_code).strip())
            except Course.DoesNotExist:
                errors.append(f"Row {index + 1}: Course with code '{course_code}' does not exist.")
                continue

            department = class_assigned.department
            semester = class_assigned.semester
            
            # Check if the course is offered in the given semester and department
            try:
                CourseOffering.objects.get(
                    course=course,
                    department=department,
                    semester=semester
                )
            except CourseOffering.DoesNotExist:
                errors.append(f"Row {index + 1}: Course {course.short_name} not offered in {department.name} department in {semester.name} semester.")
                continue

            teacher = None
            if pd.notna(teacher_name):
                try:
                    teacher = Teacher.objects.get(name__iexact=str(teacher_name).strip())
                except Teacher.DoesNotExist:
                    errors.append(f"Row {index + 1}: Teacher with name {teacher_name} does not exist.")
                    # Here we continue without assigning a teacher, as it's optional

            # Attempt to create or update the CourseAssignment
            try:
                course_assignment, created = CourseAssignment.objects.update_or_create(
                    class_assigned=class_assigned,
                    course=course,
                    defaults={'teacher': teacher}
                )
                if not created:
                    errors.append(f"Row {index + 1}: Course assignment for Class {class_name}, Course {course_code} already exists.")
            except Exception as e:
                errors.append(f"Row {index + 1}: An error occurred while processing the assignment: {str(e)}")

    # Render the template with any accumulated errors
    return render(request, 'timetable/upload_course_assignments_excel.html', {'errors': errors})

def upload_schedule(request):
    errors = []  # List to collect error messages

    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        
        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            errors.append(f"Invalid Excel file format: {e}")
            return render(request, 'timetable/upload_schedule.html', {'errors': errors})

        # Loop through each row in the Excel sheet
        for index, row in df.iterrows():
            class_name = row.get('Class_Name')
            section_name = row.get('Class_Section')
            semester_name = row.get('Class_semster')
            course_code = row.get('Course_code')
            day_name = row.get('Day')
            room_name = row.get('Room')
            timeslot_slot = row.get('Timeslot')  # Assuming Timeslot column holds the slot

            # Skip rows with missing required fields
            if pd.isna(class_name) or pd.isna(section_name) or pd.isna(semester_name) or pd.isna(course_code):
                continue

            # Trim and clean data
            class_name = str(class_name).strip()
            section_name = str(section_name).strip()
            semester_name = str(semester_name).strip()
            course_code = str(course_code).strip()
            day_name = str(day_name).strip()
            room_name = str(room_name).strip()
            timeslot_slot = str(timeslot_slot).strip()

            # Get the Class object
            try:
                class_assigned = Class.objects.get(
                    name__iexact=class_name,
                    section__iexact=section_name,
                    semester__name__iexact=semester_name
                )
            except Class.DoesNotExist:
                errors.append(f"Row {index + 1}: Class '{class_name}' with section '{section_name}' and semester '{semester_name}' does not exist.")
                continue

            # Get the CourseAssignment object
            try:
                course_assignment = CourseAssignment.objects.get(
                    class_assigned=class_assigned,
                    course__course_code=course_code
                )
            except CourseAssignment.DoesNotExist:
                errors.append(f"Row {index + 1}: Course assignment for class '{class_name}' and course '{course_code}' does not exist.")
                continue

            # Get the Day object
            try:
                day = Day.objects.get(name__iexact=day_name)
            except Day.DoesNotExist:
                errors.append(f"Row {index + 1}: Day '{day_name}' does not exist.")
                continue

            # Get the Room object
            try:
                room = Room.objects.get(name__iexact=room_name)
            except Room.DoesNotExist:
                errors.append(f"Row {index + 1}: Room '{room_name}' does not exist.")
                continue

            # Get the Timeslot object
            try:
                timeslot = Timeslot.objects.get(slot__iexact=timeslot_slot)
            except Timeslot.DoesNotExist:
                errors.append(f"Row {index + 1}: Timeslot '{timeslot_slot}' does not exist.")
                continue

            # Check room, teacher, and class availability
            if Schedule.objects.filter(day=day, room=room, timeslot=timeslot).exists():
                errors.append(f"Row {index + 1}: This room is already booked for the selected timeslot.")
                continue
            elif Schedule.objects.filter(day=day, timeslot=timeslot, course_assignment__teacher=course_assignment.teacher).exclude(course_assignment__course=course_assignment.course).exists():
                errors.append(f"Row {index + 1}: This teacher is already assigned to another course at this time.")
                continue
            elif Schedule.objects.filter(day=day, timeslot=timeslot, course_assignment__class_assigned=course_assignment.class_assigned).exists():
                errors.append(f"Row {index + 1}: This class is already assigned to a different room at this time.")
                continue

            # Create or update the Schedule
            try:
                schedule, created = Schedule.objects.update_or_create(
                    day=day,
                    room=room,
                    timeslot=timeslot,
                    defaults={'course_assignment': course_assignment}
                )
                if not created:
                    errors.append(f"Row {index + 1}: Schedule already exists for the day '{day_name}', room '{room_name}', and timeslot '{timeslot_slot}'.")
            except Exception as e:
                errors.append(f"Row {index + 1}: An error occurred while creating the schedule: {str(e)}")

    # Render the upload page with any errors
    return render(request, 'timetable/upload_schedule.html', {'errors': errors})

def upload_days_rooms_timeslots(request):
    errors = []  # List to collect error messages

    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']

        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            errors.append(f"Invalid Excel file format: {e}")
            return render(request, 'timetable/upload_days_rooms_timeslots.html', {'errors': errors})

        for index, row in df.iterrows():
            room_name = row.get('Room')
            day_name = row.get('Day')
            timeslot_slot = row.get('Timeslot')
            shift_name = row.get('Shift')  # Change variable to shift_name

            # Skip rows with missing required fields
            if pd.isna(room_name) or pd.isna(day_name) or pd.isna(timeslot_slot) or pd.isna(shift_name):
                errors.append(f"Row {index + 1}: Missing required fields.")
                continue

            # Trim and clean data
            room_name = str(room_name).strip()
            day_name = str(day_name).strip()
            timeslot_slot = str(timeslot_slot).strip()
            shift_name = str(shift_name).strip()

            # Handle Shift
            if not Shift.objects.filter(name__iexact=shift_name).exists():
                Shift.objects.create(name=shift_name)
            else:
                pass

            # Handle Room
            if not Room.objects.filter(name__iexact=room_name).exists():
                Room.objects.create(name=room_name)
            else:
                errors.append(f"Row {index + 1}: Room '{room_name}' already exists.")

            # Handle Day
            if not Day.objects.filter(name__iexact=day_name).exists():
                Day.objects.create(name=day_name)
            else:
                errors.append(f"Row {index + 1}: Day '{day_name}' already exists.")

            # Handle Timeslot
            if not Timeslot.objects.filter(slot__iexact=timeslot_slot, shift__name__iexact=shift_name).exists():
                # Get the Shift object
                try:
                    shift_obj = Shift.objects.get(name__iexact=shift_name)
                    Timeslot.objects.create(slot=timeslot_slot, shift=shift_obj)
                except Shift.DoesNotExist:
                    pass
            else:
                errors.append(f"Row {index + 1}: Timeslot '{timeslot_slot}' with shift '{shift_name}' already exists.")
    # Render the upload page with any errors
    return render(request, 'timetable/upload_days_rooms_timeslots.html', {'errors': errors})



def manage_days_rooms_timeslots(request):
    # Ensure default shifts exist
    default_shifts = ['M', 'E']
    for shift_name in default_shifts:
        Shift.objects.get_or_create(name=shift_name)
    
    if request.method == 'POST':
        item_type = request.POST.get('item_type')

        if item_type == 'bulk_delete':
            # Handle bulk delete
            room_ids = request.POST.get('room_ids').split(',') if request.POST.get('room_ids') else []
            day_ids = request.POST.get('day_ids').split(',') if request.POST.get('day_ids') else []
            timeslot_ids = request.POST.get('timeslot_ids').split(',') if request.POST.get('timeslot_ids') else []
            shift_ids = request.POST.get('shift_ids').split(',') if request.POST.get('shift_ids') else []

            Room.objects.filter(id__in=room_ids).delete()
            Day.objects.filter(id__in=day_ids).delete()
            Timeslot.objects.filter(id__in=timeslot_ids).delete()
            Shift.objects.filter(id__in=shift_ids).delete()
            return redirect('manage_data')

        elif item_type == 'add_room':
            form = addroom(request.POST)
            if form.is_valid():
                # Prevent duplicates
                if not Room.objects.filter(name=form.cleaned_data['name']).exists():
                    form.save()
                return redirect('manage_data')

        elif item_type == 'add_day':
            if Day.objects.count() >= 5:
                error_message = "You can't add more than 5 days."
                return render(request, 'timetable/manage_days_rooms_timeslots.html', {'error_message': error_message})

            form = addday(request.POST)
            if form.is_valid():
                # Prevent duplicates
                if not Day.objects.filter(name=form.cleaned_data['name']).exists():
                    form.save()
                return redirect('manage_data')

        elif item_type == 'add_timeslot':
            form = addslot(request.POST)
            if form.is_valid():
                # Prevent duplicates
                if not Timeslot.objects.filter(slot=form.cleaned_data['slot'], shift=form.cleaned_data['shift']).exists():
                    form.save()
                return redirect('manage_data')

        elif item_type == 'add_shift':
            shift_form = AddShiftForm(request.POST)
            if shift_form.is_valid():
                # Prevent duplicate shifts
                if not Shift.objects.filter(name=shift_form.cleaned_data['name']).exists():
                    shift_form.save()
                return redirect('manage_data')

    else:
        # Initialize forms for adding
        room_form = addroom()
        day_form = addday()
        timeslot_form = addslot()
        shift_form = AddShiftForm()

    rooms = Room.objects.all()
    days = Day.objects.all()
    timeslots = Timeslot.objects.all()
    shifts = Shift.objects.all()

    return render(request, 'timetable/manage_days_rooms_timeslots.html', {
        'rooms': rooms,
        'days': days,
        'timeslots': timeslots,
        'shifts': shifts,
        'room_form': room_form,
        'day_form': day_form,
        'timeslot_form': timeslot_form,
        'shift_form': shift_form,
    })

# from .models import Timetable, CourseAssignment, GeneticAlgorithmParams, Schedule
# from .genetic_algorithm import GeneticAlgorithm
# import random
# from django.db import transaction

# def get_required_slots(course_assignment):
#     """ Calculate the number of required slots based on credit hours. """
#     credit_hours = course_assignment.credit_hours.split('-')
#     lecture_hours = int(credit_hours[0])
#     lab_hours = int(credit_hours[1])

#     required_slots = []
    
#     # Calculate lecture slots
#     if lecture_hours > 0:
#         required_slots.extend([1] * lecture_hours)  # 1 slot for each hour of lecture

#     # Calculate lab slots (requirement for consecutive slots)
#     if lab_hours > 0:
#         required_slots.extend([3] * lab_hours)  # 1 lab requires 3 consecutive slots

#     return required_slots

# def satisfies_constraints(timetable):
#     """ Check if the timetable satisfies all constraints. """
#     used_slots = set()
#     class_days = {}
#     teacher_conflicts = {}
#     room_conflicts = {}

#     for course_assignment, day, room, timeslot in timetable:
#         # Check for shift constraints
#         if course_assignment.class_assigned.shift != 'M' and timeslot not in ['E', 'M2']:  # Example for shift M
#             return False
        
#         slot_id = (day, room, timeslot)

#         # Check for room conflicts
#         if slot_id in used_slots:
#             return False
#         used_slots.add(slot_id)

#         # Check class conflicts
#         if day not in class_days:
#             class_days[day] = {}
#         if course_assignment.id in class_days[day]:
#             return False
#         class_days[day][course_assignment.id] = slot_id

#         # Check teacher conflicts
#         if course_assignment.teacher_id not in teacher_conflicts:
#             teacher_conflicts[course_assignment.teacher_id] = set()
#         if slot_id in teacher_conflicts[course_assignment.teacher_id]:
#             return False
#         teacher_conflicts[course_assignment.teacher_id].add(slot_id)

#         # Check room conflicts
#         if room not in room_conflicts:
#             room_conflicts[room] = {}
#         if day not in room_conflicts[room]:
#             room_conflicts[room][day] = set()
#         if timeslot in room_conflicts[room][day]:
#             return False
#         room_conflicts[room][day].add(timeslot)

#     return True

# def initialize_population(population_size, course_assignments, days, rooms, timeslots):
#     population = []
#     for _ in range(population_size):
#         timetable = []
#         for course_assignment in course_assignments:
#             day = random.choice(days)
#             room = random.choice(rooms)
#             timeslot = random.choice(timeslots)
#             timetable.append((course_assignment, day, room, timeslot))
        
#         # Only add valid timetables
#         if satisfies_constraints(timetable):
#             population.append(timetable)
#     return population

# def fitness_function(timetable):
#     fitness = 0
#     used_slots = set()

#     for course_assignment, day, room, timeslot in timetable:
#         slot_id = (day, room, timeslot)
#         if slot_id not in used_slots:
#             used_slots.add(slot_id)
#             fitness += 1  # Increment for valid assignments
#         else:
#             fitness -= 1  # Decrement for collisions

#         # Add more fitness penalties or bonuses based on other constraints if needed

#     return fitness

# def selection(population):
#     sorted_population = sorted(population, key=fitness_function, reverse=True)
#     return sorted_population[:len(sorted_population) // 2]

# def crossover(parent1, parent2):
#     crossover_point = random.randint(1, len(parent1) - 1)
#     child = parent1[:crossover_point] + parent2[crossover_point:]
    
#     # Ensure the child timetable satisfies constraints
#     while not satisfies_constraints(child):
#         # Resample until valid
#         child = initialize_population(1, [ca for ca in parent1 + parent2], days, rooms, timeslots)[0]
    
#     return child

# def mutation(timetable, days, rooms, timeslots, mutation_rate):
#     for i in range(len(timetable)):
#         if random.random() < mutation_rate:
#             # Resample until valid
#             new_day = random.choice(days)
#             new_room = random.choice(rooms)
#             new_timeslot = random.choice(timeslots)
#             timetable[i] = (timetable[i][0], new_day, new_room, new_timeslot)
    
#     # Check if the mutated timetable satisfies constraints
#     while not satisfies_constraints(timetable):
#         timetable = initialize_population(1, [ca for ca in timetable], days, rooms, timeslots)[0]

#     return timetable

# def generate_timetable(request):
#     course_assignments = list(CourseAssignment.objects.all())
#     days = list(Day.objects.all())
#     rooms = list(Room.objects.all())
#     timeslots = list(Timeslot.objects.all())
    
#     # Check if any of the lists are empty
#     if not course_assignments or not days or not rooms or not timeslots:
#         return render(request, 'timetable/error.html', {'message': 'One or more required entities are missing. Please ensure all data is populated.'})

#     population_size = 50
#     mutation_rate = 0.1
#     max_generations = 100

#     population = initialize_population(population_size, course_assignments, days, rooms, timeslots)

#     best_timetable = None
#     best_fitness = float('-inf')

#     for generation in range(max_generations):
#         population = selection(population)
#         new_population = []

#         while len(new_population) < population_size:
#             parent1 = random.choice(population)
#             parent2 = random.choice(population)
#             child = crossover(parent1, parent2)
#             child = mutation(child, days, rooms, timeslots, mutation_rate)
#             new_population.append(child)

#         population = new_population
        
#         current_best = max(population, key=fitness_function)
#         current_best_fitness = fitness_function(current_best)

#         if current_best_fitness > best_fitness:
#             best_fitness = current_best_fitness
#             best_timetable = current_best

#         print(f'Generation {generation}: Best Fitness = {best_fitness}')

#     if best_timetable:
#         save_schedule(best_timetable)
#         return render(request, 'timetable/result.html', {'best_timetable': best_timetable})
#     else:
#         return render(request, 'timetable/error.html', {'message': 'No valid schedule found after maximum generations.'})
# def save_schedule(timetable):
#     with transaction.atomic():
#         for course_assignment, day, room, timeslot in timetable:
#             Schedule.objects.create(
#                 course_assignment=course_assignment,
#                 day=day,
#                 room=room,
#                 timeslot=timeslot,
#             )
from .ga_schedule import genetic_algorithm_schedule
from .models import Schedule

from django.db import IntegrityError

def generate_timetable(request):
    if request.method == "POST":
        best_schedule = genetic_algorithm_schedule()
        error_messages = []

        if best_schedule:
            for day, room, timeslot, course_assignment in best_schedule:
                print(f"Trying to save schedule for Day: {day}, Room: {room}, Timeslot: {timeslot}, Course Assignment: {course_assignment}")

                if Schedule.objects.filter(day=day, room=room, timeslot=timeslot, course_assignment=course_assignment).exists():
                    error_messages.append(f"Schedule for Day: {day}, Room: {room}, Timeslot: {timeslot} already exists.")
                else:
                    try:
                        Schedule.objects.create(
                            day=day,
                            room=room,
                            timeslot=timeslot,
                            course_assignment=course_assignment
                        )
                        print(f"Successfully saved schedule for Day: {day}, Room: {room}, Timeslot: {timeslot}.")
                    except IntegrityError as e:
                        error_messages.append(f"Failed to save schedule: {e}")

            return render(request, 'timetable/generate_timetable.html', {'error_messages': error_messages})

    return render(request, 'timetable/generate_timetable.html')
